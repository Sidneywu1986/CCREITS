"""
Admin Funds Routes
"""

from fastapi import APIRouter, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from typing import Optional
import asyncpg

from ..utils import DB_URL, get_admin_user, sign_cookie, verify_cookie, sql_placeholders

router = APIRouter()
api_router = APIRouter()

@router.get("/list", response_class=HTMLResponse)
async def funds_list(request: Request, page: int = 1, limit: int = 20,
                     search: str = "", exchange: str = "", status: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin

    # 搜索和筛选功能
    query = FundAdmin.all()

    # 搜索（基金代码、名称）
    if search:
        query = query.filter(
            FundAdmin.fund_code.contains(search) |
            FundAdmin.fund_name.contains(search)
        )

    # 交易所筛选
    if exchange:
        query = query.filter(FundAdmin.exchange == exchange)

    # 状态筛选
    if status:
        query = query.filter(FundAdmin.status == status)

    offset = (page - 1) * limit
    funds = await query.offset(offset).limit(limit)
    total = await query.count()

    rows = ""
    for f in funds:
        rows += f"""
        <tr>
            <td><input type="checkbox" name="fund_ids" value="{f.id}" class="batch-checkbox"></td>
            <td>{f.id}</td>
            <td>{f.fund_code}</td>
            <td>{f.fund_name}</td>
            <td>{f.exchange or '-'}</td>
            <td>{f.manager or '-'}</td>
            <td>{f.status or '-'}</td>
            <td>
                <a href="/admin/funds/edit/{f.id}" class="btn btn-sm">编辑</a>
                <a href="/admin/funds/delete/{f.id}" class="btn btn-sm btn-danger" onclick="return confirm('确定删除?')">删除</a>
            </td>
        </tr>
        """

    # 搜索和筛选表单
    search_form = f"""
    <div class="search-box">
        <form method="GET" action="/admin/funds/list">
            <input type="text" name="search" placeholder="搜索基金代码或名称" value="{search}">
            <select name="exchange">
                <option value="">所有交易所</option>
                <option value="SH" {'selected' if exchange == 'SH' else ''}>上海证券交易所</option>
                <option value="SZ" {'selected' if exchange == 'SZ' else ''}>深圳证券交易所</option>
            </select>
            <select name="status">
                <option value="">所有状态</option>
                <option value="listed" {'selected' if status == 'listed' else ''}>已上市</option>
                <option value="pending" {'selected' if status == 'pending' else ''}>待上市</option>
                <option value="delisted" {'selected' if status == 'delisted' else ''}>已退市</option>
            </select>
            <button type="submit" class="btn btn-primary">搜索</button>
            <a href="/admin/funds/list" class="btn">清空</a>
        </form>
    </div>
    """

    # 批量操作和导入导出
    batch_actions = """
    <div class="batch-actions">
        <button type="button" class="btn btn-info" onclick="batchDelete()">批量删除</button>
        <button type="button" class="btn btn-warning" onclick="batchUpdateStatus()">批量更新状态</button>
        <a href="/admin/funds/export" class="btn btn-success">导出Excel</a>
        <a href="/admin/funds/import" class="btn btn-primary">导入Excel</a>
        <a href="/admin/funds/template" class="btn">下载模板</a>
    </div>
    """

    html = render_admin_page("基金管理", f"""
        <div class="page-header">
            <h2>基金管理</h2>
            <a href="/admin/funds/create" class="btn btn-primary">+ 新增基金</a>
        </div>
        {search_form}
        {batch_actions}
        <table class="data-table">
            <thead>
                <tr>
                    <th><input type="checkbox" onclick="toggleBatch(this)"></th>
                    <th>ID</th>
                    <th>基金代码</th>
                    <th>基金名称</th>
                    <th>交易所</th>
                    <th>管理人</th>
                    <th>状态</th>
                    <th>操作</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        <div class="pagination">
            <span>共 {total} 条记录</span>
        </div>
        <script>
        function toggleBatch(source) {{
            var checkboxes = document.querySelectorAll('.batch-checkbox');
            for (var i = 0; i < checkboxes.length; i++) {{
                checkboxes[i].checked = source.checked;
            }}
        }}

        function batchDelete() {{
            var selected = document.querySelectorAll('.batch-checkbox:checked');
            if (selected.length === 0) {{
                alert('请选择要删除的基金');
                return;
            }}
            if (confirm('确定删除选中的 ' + selected.length + ' 只基金吗?')) {{
                var form = document.createElement('form');
                form.method = 'POST';
                form.action = '/admin/funds/batch-delete';
                form.innerHTML = '<input type="hidden" name="fund_ids" value="' + Array.from(selected).map(cb => cb.value).join(',') + '">';
                document.body.appendChild(form);
                form.submit();
            }}
        }}

        function batchUpdateStatus() {{
            var selected = document.querySelectorAll('.batch-checkbox:checked');
            if (selected.length === 0) {{
                alert('请选择要更新的基金');
                return;
            }}
            var status = prompt('请输入新状态 (listed/pending/delisted):');
            if (status) {{
                var form = document.createElement('form');
                form.method = 'POST';
                form.action = '/admin/funds/batch-update';
                form.innerHTML = '<input type="hidden" name="fund_ids" value="' + Array.from(selected).map(cb => cb.value).join(',') + '"><input type="hidden" name="status" value="' + status + '">';
                document.body.appendChild(form);
                form.submit();
            }}
        }}
        </script>
    """, user)
    return HTMLResponse(content=html)


# ========== 基金创建页面 ==========
@router.get("/create", response_class=HTMLResponse)
async def fund_create_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("创建基金", f"""
        <div class="page-header">
            <h2>创建基金</h2>
            <a href="/admin/funds/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/funds/create" class="data-form">
            <div class="form-row">
                <div class="form-group">
                    <label>基金代码 *</label>
                    <input type="text" name="fund_code" required placeholder="例如: 508000" maxlength="10">
                </div>
                <div class="form-group">
                    <label>基金名称 *</label>
                    <input type="text" name="fund_name" required placeholder="基金简称" maxlength="100">
                </div>
            </div>
            <div class="form-group">
                <label>完整名称</label>
                <input type="text" name="full_name" placeholder="基金完整名称" maxlength="200">
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>交易所</label>
                    <select name="exchange">
                        <option value="">请选择</option>
                        <option value="SH">上海证券交易所</option>
                        <option value="SZ">深圳证券交易所</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>状态</label>
                    <select name="status">
                        <option value="listed">已上市</option>
                        <option value="pending">待上市</option>
                        <option value="delisted">已退市</option>
                    </select>
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>IPO日期</label>
                    <input type="text" name="ipo_date" placeholder="YYYY-MM-DD">
                </div>
                <div class="form-group">
                    <label>IPO价格</label>
                    <input type="number" name="ipo_price" step="0.01" placeholder="0.00">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>总份额（亿份）</label>
                    <input type="number" name="total_shares" step="0.01" placeholder="0.00">
                </div>
                <div class="form-group">
                    <label>净值</label>
                    <input type="number" name="nav" step="0.0001" placeholder="0.0000">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>管理人</label>
                    <input type="text" name="manager" placeholder="基金管理公司" maxlength="100">
                </div>
                <div class="form-group">
                    <label>托管人</label>
                    <input type="text" name="custodian" placeholder="基金托管银行" maxlength="100">
                </div>
            </div>
            <div class="form-group">
                <label>资产类型</label>
                <input type="text" name="asset_type" placeholder="例如: 产业园、高速公路、仓储物流" maxlength="50">
            </div>
            <div class="form-group">
                <label>底层资产描述</label>
                <textarea name="underlying_assets" rows="4" placeholder="描述基金持有的底层资产情况"></textarea>
            </div>
            <button type="submit" class="btn btn-primary">提交</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 创建基金提交
@router.post("/create")
async def fund_create_submit(
    fund_code: str = Form(...),
    fund_name: str = Form(...),
    full_name: str = Form(""),
    exchange: str = Form(""),
    ipo_date: str = Form(""),
    ipo_price: str = Form(""),
    total_shares: str = Form(""),
    nav: str = Form(""),
    manager: str = Form(""),
    custodian: str = Form(""),
    asset_type: str = Form(""),
    underlying_assets: str = Form(""),
    status: str = Form("listed")
):
    from admin_models import FundAdmin

    await FundAdmin.create(
        fund_code=fund_code,
        fund_name=fund_name,
        full_name=full_name or None,
        exchange=exchange or None,
        ipo_date=ipo_date or None,
        ipo_price=float(ipo_price) if ipo_price else None,
        total_shares=float(total_shares) if total_shares else None,
        nav=float(nav) if nav else None,
        manager=manager or None,
        custodian=custodian or None,
        asset_type=asset_type or None,
        underlying_assets=underlying_assets or None,
        status=status or None
    )

    return RedirectResponse(url="/admin/funds/list", status_code=302)


# ========== 基金编辑页面 ==========
@router.get("/edit/{fund_id}", response_class=HTMLResponse)
async def fund_edit_page(request: Request, fund_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    fund = await FundAdmin.filter(id=fund_id).first()

    if not fund:
        return HTMLResponse(content="<script>alert('基金不存在');history.back();</script>")

    html = render_admin_page("编辑基金", f"""
        <div class="page-header">
            <h2>编辑基金</h2>
            <a href="/admin/funds/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/funds/edit/{fund.id}" class="data-form">
            <div class="form-row">
                <div class="form-group">
                    <label>基金代码 *</label>
                    <input type="text" name="fund_code" value="{fund.fund_code}" required maxlength="10">
                </div>
                <div class="form-group">
                    <label>基金名称 *</label>
                    <input type="text" name="fund_name" value="{fund.fund_name}" required maxlength="100">
                </div>
            </div>
            <div class="form-group">
                <label>完整名称</label>
                <input type="text" name="full_name" value="{fund.full_name or ''}" maxlength="200">
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>交易所</label>
                    <select name="exchange">
                        <option value="" {'selected' if not fund.exchange else ''}>请选择</option>
                        <option value="SH" {'selected' if fund.exchange == 'SH' else ''}>上海证券交易所</option>
                        <option value="SZ" {'selected' if fund.exchange == 'SZ' else ''}>深圳证券交易所</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>状态</label>
                    <select name="status">
                        <option value="listed" {'selected' if fund.status == 'listed' else ''}>已上市</option>
                        <option value="pending" {'selected' if fund.status == 'pending' else ''}>待上市</option>
                        <option value="delisted" {'selected' if fund.status == 'delisted' else ''}>已退市</option>
                    </select>
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>IPO日期</label>
                    <input type="text" name="ipo_date" value="{fund.ipo_date or ''}" placeholder="YYYY-MM-DD">
                </div>
                <div class="form-group">
                    <label>IPO价格</label>
                    <input type="number" name="ipo_price" step="0.01" value="{fund.ipo_price or ''}">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>总份额（亿份）</label>
                    <input type="number" name="total_shares" step="0.01" value="{fund.total_shares or ''}">
                </div>
                <div class="form-group">
                    <label>净值</label>
                    <input type="number" name="nav" step="0.0001" value="{fund.nav or ''}">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>管理人</label>
                    <input type="text" name="manager" value="{fund.manager or ''}" maxlength="100">
                </div>
                <div class="form-group">
                    <label>托管人</label>
                    <input type="text" name="custodian" value="{fund.custodian or ''}" maxlength="100">
                </div>
            </div>
            <div class="form-group">
                <label>资产类型</label>
                <input type="text" name="asset_type" value="{fund.asset_type or ''}" maxlength="50">
            </div>
            <div class="form-group">
                <label>底层资产描述</label>
                <textarea name="underlying_assets" rows="4">{fund.underlying_assets or ''}</textarea>
            </div>
            <button type="submit" class="btn btn-primary">保存修改</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 编辑基金提交
@router.post("/edit/{fund_id}")
async def fund_edit_submit(
    fund_id: int,
    fund_code: str = Form(...),
    fund_name: str = Form(...),
    full_name: str = Form(""),
    exchange: str = Form(""),
    ipo_date: str = Form(""),
    ipo_price: str = Form(""),
    total_shares: str = Form(""),
    nav: str = Form(""),
    manager: str = Form(""),
    custodian: str = Form(""),
    asset_type: str = Form(""),
    underlying_assets: str = Form(""),
    status: str = Form("listed")
):
    from admin_models import FundAdmin
    fund = await FundAdmin.filter(id=fund_id).first()

    if not fund:
        return HTMLResponse(content="<script>alert('基金不存在');history.back();</script>")

    fund.fund_code = fund_code
    fund.fund_name = fund_name
    fund.full_name = full_name or None
    fund.exchange = exchange or None
    fund.ipo_date = ipo_date or None
    fund.ipo_price = float(ipo_price) if ipo_price else None
    fund.total_shares = float(total_shares) if total_shares else None
    fund.nav = float(nav) if nav else None
    fund.manager = manager or None
    fund.custodian = custodian or None
    fund.asset_type = asset_type or None
    fund.underlying_assets = underlying_assets or None
    fund.status = status or None
    await fund.save()

    return RedirectResponse(url="/admin/funds/list", status_code=302)


# 基金根路径重定向到列表
@router.get("/admin/funds")
async def funds_redirect():
    return RedirectResponse(url="/admin/funds/list", status_code=302)


# 删除基金
@router.get("/delete/{fund_id}")
async def fund_delete(request: Request, fund_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    fund = await FundAdmin.filter(id=fund_id).first()

    if not fund:
        return HTMLResponse(content="<script>alert('基金不存在');history.back();</script>")

    await fund.delete()
    return RedirectResponse(url="/admin/funds/list", status_code=302)


# ========== 用户列表页面 ==========
@router.get("/admin/users/list", response_class=HTMLResponse)
async def users_list(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")
    
    from admin_models import UserAdmin
    users = await UserAdmin.all().limit(50)
    
    rows = ""
    for u in users:
        status = '<span class="badge badge-success">启用</span>' if u.is_active else '<span class="badge badge-danger">禁用</span>'
        rows += f"""
        <tr>
            <td>{u.id}</td>
            <td>{u.username}</td>
            <td>{u.email}</td>
            <td>{status}</td>
            <td>{str(u.created_at)[:19]}</td>
            <td>
                <a href="/admin/users/edit/{u.id}" class="btn btn-sm">编辑</a>
            </td>
        </tr>
        """
    
    html = render_admin_page("用户管理", f"""
        <div class="page-header">
            <h2>用户管理</h2>
            <a href="/admin/users/create" class="btn btn-primary">+ 新增用户</a>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>用户名</th>
                    <th>邮箱</th>
                    <th>状态</th>
                    <th>创建时间</th>
                    <th>操作</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
    """, user)
    return HTMLResponse(content=html)


# ========== 用户创建页面 ==========
@router.get("/admin/users/create", response_class=HTMLResponse)
async def user_create_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("创建用户", f"""
        <div class="page-header">
            <h2>创建用户</h2>
            <a href="/admin/users/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/users/create" class="data-form">
            <div class="form-group">
                <label>用户名 *</label>
                <input type="text" name="username" required placeholder="登录用户名" maxlength="50">
            </div>
            <div class="form-group">
                <label>邮箱 *</label>
                <input type="email" name="email" required placeholder="user@example.com" maxlength="100">
            </div>
            <div class="form-group">
                <label>密码 *</label>
                <input type="password" name="password" required placeholder="设置登录密码" minlength="6">
                <div class="help-text">密码长度至少6位</div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>状态</label>
                    <select name="is_active">
                        <option value="1">启用</option>
                        <option value="0">禁用</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>超级用户</label>
                    <select name="is_superuser">
                        <option value="0">否</option>
                        <option value="1">是</option>
                    </select>
                </div>
            </div>
            <button type="submit" class="btn btn-primary">提交</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 创建用户提交
@router.post("/admin/users/create")
async def user_create_submit(
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    is_active: str = Form("1"),
    is_superuser: str = Form("0")
):
    from admin_models import UserAdmin
    from passlib.hash import bcrypt

    # 检查用户名是否已存在
    existing = await UserAdmin.filter(username=username).first()
    if existing:
        return HTMLResponse(content="<script>alert('用户名已存在');history.back();</script>")

    # 检查邮箱是否已存在
    existing_email = await UserAdmin.filter(email=email).first()
    if existing_email:
        return HTMLResponse(content="<script>alert('邮箱已被使用');history.back();</script>")

    await UserAdmin.create(
        username=username,
        email=email,
        password_hash=bcrypt.hash(password),
        is_active=is_active == "1",
        is_superuser=is_superuser == "1"
    )

    return RedirectResponse(url="/admin/users/list", status_code=302)


# ========== 用户编辑页面 ==========
@router.get("/admin/users/edit/{user_id}", response_class=HTMLResponse)
async def user_edit_page(request: Request, user_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import UserAdmin
    u = await UserAdmin.filter(id=user_id).first()

    if not u:
        return HTMLResponse(content="<script>alert('用户不存在');history.back();</script>")

    html = render_admin_page("编辑用户", f"""
        <div class="page-header">
            <h2>编辑用户</h2>
            <a href="/admin/users/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/users/edit/{u.id}" class="data-form">
            <div class="form-group">
                <label>用户名 *</label>
                <input type="text" name="username" value="{u.username}" required maxlength="50">
            </div>
            <div class="form-group">
                <label>邮箱 *</label>
                <input type="email" name="email" value="{u.email}" required maxlength="100">
            </div>
            <div class="form-group">
                <label>新密码（留空则不修改）</label>
                <input type="password" name="password" placeholder="不修改请留空" minlength="6">
                <div class="help-text">如需修改密码，输入新密码（至少6位）</div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label>状态</label>
                    <select name="is_active">
                        <option value="1" {'selected' if u.is_active else ''}>启用</option>
                        <option value="0" {'selected' if not u.is_active else ''}>禁用</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>超级用户</label>
                    <select name="is_superuser">
                        <option value="1" {'selected' if u.is_superuser else ''}>是</option>
                        <option value="0" {'selected' if not u.is_superuser else ''}>否</option>
                    </select>
                </div>
            </div>
            <button type="submit" class="btn btn-primary">保存修改</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 编辑用户提交
@router.post("/admin/users/edit/{user_id}")
async def user_edit_submit(
    user_id: int,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(""),
    is_active: str = Form("1"),
    is_superuser: str = Form("0")
):
    from admin_models import UserAdmin
    from passlib.hash import bcrypt
    u = await UserAdmin.filter(id=user_id).first()

    if not u:
        return HTMLResponse(content="<script>alert('用户不存在');history.back();</script>")

    # 检查用户名是否被其他用户占用
    existing = await UserAdmin.filter(username=username).exclude(id=user_id).first()
    if existing:
        return HTMLResponse(content="<script>alert('用户名已被其他用户使用');history.back();</script>")

    # 检查邮箱是否被其他用户占用
    existing_email = await UserAdmin.filter(email=email).exclude(id=user_id).first()
    if existing_email:
        return HTMLResponse(content="<script>alert('邮箱已被其他用户使用');history.back();</script>")

    u.username = username
    u.email = email
    u.is_active = is_active == "1"
    u.is_superuser = is_superuser == "1"

    if password:
        u.password_hash = bcrypt.hash(password)

    await u.save()

    return RedirectResponse(url="/admin/users/list", status_code=302)


def render_admin_page(title, content, username):
    """渲染管理页面框架"""
    return f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>REITs Admin - {title}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f5f7fa;
        }}
        .sidebar {{
            position: fixed;
            left: 0; top: 0;
            width: 220px;
            height: 100vh;
            background: #1a1a2e;
            color: white;
        }}
        .sidebar-header {{
            padding: 20px;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            font-size: 18px;
            font-weight: bold;
        }}
        .nav-item {{
            display: block;
            padding: 15px 20px;
            color: rgba(255,255,255,0.8);
            text-decoration: none;
            border-left: 3px solid transparent;
            transition: all 0.3s;
        }}
        .nav-item:hover, .nav-item.active {{
            background: rgba(255,255,255,0.1);
            border-left-color: #667eea;
            color: white;
        }}
        .main {{
            margin-left: 220px;
            padding: 30px;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 1px solid #eee;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
        }}
        .logout {{
            color: #e74c3c;
            text-decoration: none;
            font-size: 14px;
        }}
        .page-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 25px;
        }}
        .btn {{
            padding: 8px 16px;
            border-radius: 6px;
            text-decoration: none;
            font-size: 14px;
            cursor: pointer;
            border: none;
            display: inline-block;
        }}
        .btn-primary {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        .btn-sm {{ padding: 5px 12px; font-size: 12px; }}
        .btn-danger {{ background: #e74c3c; color: white; }}
        .data-table {{
            width: 100%;
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
            border-collapse: collapse;
        }}
        .data-table th, .data-table td {{
            padding: 15px;
            text-align: left;
            border-bottom: 1px solid #eee;
        }}
        .data-table th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #333;
        }}
        .data-table tr:hover {{ background: #f8f9fa; }}
        .badge {{
            padding: 4px 10px;
            border-radius: 4px;
            font-size: 12px;
        }}
        .badge-success {{ background: #d4edda; color: #155724; }}
        .badge-danger {{ background: #f8d7da; color: #721c24; }}
        .pagination {{
            margin-top: 20px;
            text-align: right;
            color: #666;
        }}
        .data-form {{
            background: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
            max-width: 900px;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-group label {{
            display: block;
            margin-bottom: 8px;
            color: #555;
            font-weight: 500;
        }}
        .form-group input,
        .form-group textarea,
        .form-group select {{
            width: 100%;
            padding: 12px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 14px;
            font-family: inherit;
            background: white;
        }}
        .form-group input:focus,
        .form-group textarea:focus,
        .form-group select:focus {{
            outline: none;
            border-color: #667eea;
        }}
        .form-group textarea {{
            resize: vertical;
            min-height: 80px;
        }}
        .form-row {{
            display: flex;
            gap: 20px;
        }}
        .form-row .form-group {{
            flex: 1;
        }}
        .help-text {{
            font-size: 12px;
            color: #999;
            margin-top: 4px;
        }}
    </style>
</head>
<body>
    <div class="sidebar">
        <div class="sidebar-header">REITs 管理平台</div>
        <a href="/admin/" class="nav-item">仪表盘</a>
        <a href="/admin/users/list" class="nav-item{' active' if '用户' in title else ''}">用户管理</a>
        <a href="/admin/funds/list" class="nav-item{' active' if '基金' in title else ''}">基金管理</a>
        <a href="/admin/announcements/list" class="nav-item">公告管理</a>
        <a href="/admin/roles/list" class="nav-item">角色管理</a>
        <a href="/admin/permissions/list" class="nav-item">权限管理</a>
    </div>
    <div class="main">
        <div class="header">
            <h1>{title}</h1>
            <div class="user-info">
                <span>欢迎, {username}</span>
                <a href="/admin/logout" class="logout">退出</a>
            </div>
        </div>
        {content}
    </div>
</body>
</html>
    """


# ========== 公告管理页面 ==========

# 公告根路径重定向到列表
@router.get("/admin/announcements")
async def announcements_redirect():
    return RedirectResponse(url="/admin/announcements/list", status_code=302)


# 公告列表页面
@router.get("/admin/announcements/list", response_class=HTMLResponse)
async def announcements_list(request: Request, page: int = 1, limit: int = 20, search: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import AnnouncementAdmin

    # 搜索功能
    query = AnnouncementAdmin.all()
    if search:
        query = query.filter(
            AnnouncementAdmin.title.contains(search) |
            AnnouncementAdmin.fund_code.contains(search)
        )

    offset = (page - 1) * limit
    announcements = await query.offset(offset).limit(limit).order_by("-publish_date")
    total = await query.count()

    rows = ""
    for a in announcements:
        rows += f"""
        <tr>
            <td>{a.id}</td>
            <td>{a.fund_code}</td>
            <td>{a.title}</td>
            <td>{str(a.publish_date)[:10]}</td>
            <td>{a.announcement_type or '-'}</td>
            <td>
                <a href="/admin/announcements/detail/{a.id}" class="btn btn-sm">查看</a>
                <a href="/admin/announcements/edit/{a.id}" class="btn btn-sm">编辑</a>
                <a href="/admin/announcements/delete/{a.id}" class="btn btn-sm btn-danger" onclick="return confirm('确定删除?')">删除</a>
            </td>
        </tr>
        """

    # 搜索表单
    search_form = f"""
    <div class="search-box">
        <form method="GET" action="/admin/announcements/list">
            <input type="text" name="search" placeholder="搜索公告标题或基金代码" value="{search}">
            <button type="submit" class="btn btn-primary">搜索</button>
            <a href="/admin/announcements/list" class="btn">清空</a>
        </form>
    </div>
    """

    html = render_admin_page("公告管理", f"""
        <div class="page-header">
            <h2>公告管理</h2>
            {search_form}
            <a href="/admin/announcements/create" class="btn btn-primary">+ 新增公告</a>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>基金代码</th>
                    <th>标题</th>
                    <th>发布日期</th>
                    <th>类型</th>
                    <th>操作</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        <div class="pagination">
            <span>共 {total} 条记录</span>
        </div>
    """, user)
    return HTMLResponse(content=html)


# 创建公告页面
@router.get("/admin/announcements/create", response_class=HTMLResponse)
async def announcement_create_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("创建公告", f"""
        <div class="page-header">
            <h2>创建公告</h2>
            <a href="/admin/announcements/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/announcements/create" class="data-form">
            <div class="form-group">
                <label>基金代码 *</label>
                <input type="text" name="fund_code" required placeholder="例如: 150050">
            </div>
            <div class="form-group">
                <label>标题 *</label>
                <input type="text" name="title" required placeholder="公告标题">
            </div>
            <div class="form-group">
                <label>发布日期 *</label>
                <input type="date" name="publish_date" required value="{datetime.now().strftime('%Y-%m-%d')}">
            </div>
            <div class="form-group">
                <label>公告类型</label>
                <select name="announcement_type">
                    <option value="">请选择</option>
                    <option value="定期报告">定期报告</option>
                    <option value="季度报告">季度报告</option>
                    <option value="年度报告">年度报告</option>
                    <option value="临时公告">临时公告</option>
                    <option value="分红公告">分红公告</option>
                    <option value="其他">其他</option>
                </select>
            </div>
            <div class="form-group">
                <label>内容</label>
                <textarea name="content" rows="6" placeholder="公告内容"></textarea>
            </div>
            <div class="form-group">
                <label>PDF链接</label>
                <input type="text" name="pdf_url" placeholder="PDF文件链接">
            </div>
            <button type="submit" class="btn btn-primary">提交</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 创建公告提交
@router.post("/admin/announcements/create")
async def announcement_create_submit(
    fund_code: str = Form(...),
    title: str = Form(...),
    publish_date: str = Form(...),
    announcement_type: str = Form(""),
    content: str = Form(""),
    pdf_url: str = Form("")
):
    from admin_models import AnnouncementAdmin

    await AnnouncementAdmin.create(
        fund_code=fund_code,
        title=title,
        publish_date=publish_date + "T00:00:00",
        announcement_type=announcement_type,
        content=content,
        pdf_url=pdf_url
    )

    return RedirectResponse(url="/admin/announcements/list", status_code=302)


# 公告详情页面
@router.get("/admin/announcements/detail/{announcement_id}", response_class=HTMLResponse)
async def announcement_detail(request: Request, announcement_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import AnnouncementAdmin
    announcement = await AnnouncementAdmin.filter(id=announcement_id).first()

    if not announcement:
        return HTMLResponse(content="<script>alert('公告不存在');history.back();</script>")

    html = render_admin_page("公告详情", f"""
        <div class="page-header">
            <h2>公告详情</h2>
            <a href="/admin/announcements/list" class="btn">返回列表</a>
            <a href="/admin/announcements/edit/{announcement.id}" class="btn btn-primary">编辑</a>
        </div>
        <div class="detail-box">
            <div class="detail-row">
                <span class="detail-label">ID:</span>
                <span class="detail-value">{announcement.id}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">基金代码:</span>
                <span class="detail-value">{announcement.fund_code}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">标题:</span>
                <span class="detail-value">{announcement.title}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">发布日期:</span>
                <span class="detail-value">{str(announcement.publish_date)[:19]}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">公告类型:</span>
                <span class="detail-value">{announcement.announcement_type or '-'}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">PDF链接:</span>
                <span class="detail-value">{announcement.pdf_url or '-'}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">内容:</span>
                <span class="detail-value" style="white-space: pre-wrap;">{announcement.content or '-'}</span>
            </div>
        </div>
    """, user)
    return HTMLResponse(content=html)


# 编辑公告页面
@router.get("/admin/announcements/edit/{announcement_id}", response_class=HTMLResponse)
async def announcement_edit_page(request: Request, announcement_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import AnnouncementAdmin
    announcement = await AnnouncementAdmin.filter(id=announcement_id).first()

    if not announcement:
        return HTMLResponse(content="<script>alert('公告不存在');history.back();</script>")

    # 格式化日期
    publish_date = str(announcement.publish_date)[:10]

    html = render_admin_page("编辑公告", f"""
        <div class="page-header">
            <h2>编辑公告</h2>
            <a href="/admin/announcements/detail/{announcement.id}" class="btn">取消编辑</a>
        </div>
        <form method="POST" action="/admin/announcements/edit/{announcement.id}" class="data-form">
            <div class="form-group">
                <label>基金代码 *</label>
                <input type="text" name="fund_code" value="{announcement.fund_code}" required>
            </div>
            <div class="form-group">
                <label>标题 *</label>
                <input type="text" name="title" value="{announcement.title}" required>
            </div>
            <div class="form-group">
                <label>发布日期 *</label>
                <input type="date" name="publish_date" value="{publish_date}" required>
            </div>
            <div class="form-group">
                <label>公告类型</label>
                <select name="announcement_type">
                    <option value="">请选择</option>
                    <option value="定期报告" {'selected' if announcement.announcement_type == '定期报告' else ''}>定期报告</option>
                    <option value="季度报告" {'selected' if announcement.announcement_type == '季度报告' else ''}>季度报告</option>
                    <option value="年度报告" {'selected' if announcement.announcement_type == '年度报告' else ''}>年度报告</option>
                    <option value="临时公告" {'selected' if announcement.announcement_type == '临时公告' else ''}>临时公告</option>
                    <option value="分红公告" {'selected' if announcement.announcement_type == '分红公告' else ''}>分红公告</option>
                    <option value="其他" {'selected' if announcement.announcement_type == '其他' else ''}>其他</option>
                </select>
            </div>
            <div class="form-group">
                <label>内容</label>
                <textarea name="content" rows="6">{announcement.content or ''}</textarea>
            </div>
            <div class="form-group">
                <label>PDF链接</label>
                <input type="text" name="pdf_url" value="{announcement.pdf_url or ''}">
            </div>
            <button type="submit" class="btn btn-primary">保存修改</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 编辑公告提交
@router.post("/admin/announcements/edit/{announcement_id}")
async def announcement_edit_submit(
    announcement_id: int,
    fund_code: str = Form(...),
    title: str = Form(...),
    publish_date: str = Form(...),
    announcement_type: str = Form(""),
    content: str = Form(""),
    pdf_url: str = Form("")
):
    from admin_models import AnnouncementAdmin
    announcement = await AnnouncementAdmin.filter(id=announcement_id).first()

    if not announcement:
        return HTMLResponse(content="<script>alert('公告不存在');history.back();</script>")

    # 更新公告
    announcement.fund_code = fund_code
    announcement.title = title
    announcement.publish_date = publish_date + "T00:00:00"
    announcement.announcement_type = announcement_type
    announcement.content = content
    announcement.pdf_url = pdf_url
    await announcement.save()

    return RedirectResponse(url=f"/admin/announcements/detail/{announcement_id}", status_code=302)


# 删除公告
@router.get("/admin/announcements/delete/{announcement_id}")
async def announcement_delete(request: Request, announcement_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import AnnouncementAdmin
    announcement = await AnnouncementAdmin.filter(id=announcement_id).first()

    if not announcement:
        return HTMLResponse(content="<script>alert('公告不存在');history.back();</script>")

    await announcement.delete()
    return RedirectResponse(url="/admin/announcements/list", status_code=302)


# ========== 角色管理页面 ==========

# 角色列表页面
@router.get("/admin/roles/list", response_class=HTMLResponse)
async def roles_list(request: Request, page: int = 1, limit: int = 20, search: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import RoleAdmin, PermissionAdmin

    # 搜索功能
    query = RoleAdmin.all()
    if search:
        query = query.filter(
            RoleAdmin.name.contains(search) |
            RoleAdmin.description.contains(search)
        )

    offset = (page - 1) * limit
    roles = await query.offset(offset).limit(limit).order_by("-created_at")
    total = await query.count()

    # 获取每个角色的权限数量
    role_permissions = {}
    for role in roles:
        role_permissions[role.id] = await PermissionAdmin.filter(
            id__in=await role.permissions.all().values_list("id", flat=True)
        ).count()

    rows = ""
    for r in roles:
        rows += f"""
        <tr>
            <td>{r.id}</td>
            <td>{r.name}</td>
            <td>{r.description or '-'}</td>
            <td>{role_permissions.get(r.id, 0)} 个权限</td>
            <td>{str(r.created_at)[:19]}</td>
            <td>
                <a href="/admin/roles/edit/{r.id}" class="btn btn-sm">编辑</a>
                <a href="/admin/roles/permissions/{r.id}" class="btn btn-sm btn-info">权限分配</a>
                <a href="/admin/roles/delete/{r.id}" class="btn btn-sm btn-danger" onclick="return confirm('确定删除?')">删除</a>
            </td>
        </tr>
        """

    # 搜索表单
    search_form = f"""
    <div class="search-box">
        <form method="GET" action="/admin/roles/list">
            <input type="text" name="search" placeholder="搜索角色名称或描述" value="{search}">
            <button type="submit" class="btn btn-primary">搜索</button>
            <a href="/admin/roles/list" class="btn">清空</a>
        </form>
    </div>
    """

    html = render_admin_page("角色管理", f"""
        <div class="page-header">
            <h2>角色管理</h2>
            {search_form}
            <a href="/admin/roles/create" class="btn btn-primary">+ 新增角色</a>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>角色名称</th>
                    <th>描述</th>
                    <th>权限数量</th>
                    <th>创建时间</th>
                    <th>操作</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        <div class="pagination">
            <span>共 {total} 条记录</span>
        </div>
    """, user)
    return HTMLResponse(content=html)


# 创建角色页面
@router.get("/admin/roles/create", response_class=HTMLResponse)
async def role_create_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("创建角色", f"""
        <div class="page-header">
            <h2>创建角色</h2>
            <a href="/admin/roles/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/roles/create" class="data-form">
            <div class="form-group">
                <label>角色名称 *</label>
                <input type="text" name="name" required placeholder="例如: 编辑员">
            </div>
            <div class="form-group">
                <label>描述</label>
                <textarea name="description" rows="3" placeholder="角色描述"></textarea>
            </div>
            <button type="submit" class="btn btn-primary">提交</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 创建角色提交
@router.post("/admin/roles/create")
async def role_create_submit(
    name: str = Form(...),
    description: str = Form("")
):
    from admin_models import RoleAdmin

    await RoleAdmin.create(
        name=name,
        description=description
    )

    return RedirectResponse(url="/admin/roles/list", status_code=302)


# 编辑角色页面
@router.get("/admin/roles/edit/{role_id}", response_class=HTMLResponse)
async def role_edit_page(request: Request, role_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import RoleAdmin
    role = await RoleAdmin.filter(id=role_id).first()

    if not role:
        return HTMLResponse(content="<script>alert('角色不存在');history.back();</script>")

    html = render_admin_page("编辑角色", f"""
        <div class="page-header">
            <h2>编辑角色</h2>
            <a href="/admin/roles/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/roles/edit/{role.id}" class="data-form">
            <div class="form-group">
                <label>角色名称 *</label>
                <input type="text" name="name" value="{role.name}" required>
            </div>
            <div class="form-group">
                <label>描述</label>
                <textarea name="description" rows="3">{role.description or ''}</textarea>
            </div>
            <button type="submit" class="btn btn-primary">保存修改</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 编辑角色提交
@router.post("/admin/roles/edit/{role_id}")
async def role_edit_submit(
    role_id: int,
    name: str = Form(...),
    description: str = Form("")
):
    from admin_models import RoleAdmin
    role = await RoleAdmin.filter(id=role_id).first()

    if not role:
        return HTMLResponse(content="<script>alert('角色不存在');history.back();</script>")

    role.name = name
    role.description = description
    await role.save()

    return RedirectResponse(url="/admin/roles/list", status_code=302)


# 删除角色
@router.get("/admin/roles/delete/{role_id}")
async def role_delete(request: Request, role_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import RoleAdmin, user_roles
    role = await RoleAdmin.filter(id=role_id).first()

    if not role:
        return HTMLResponse(content="<script>alert('角色不存在');history.back();</script>")

    # 删除用户-角色关联
    await user_roles.filter(role_id=role_id).delete()
    # 删除角色-权限关联
    await role.permissions.clear()
    # 删除角色
    await role.delete()

    return RedirectResponse(url="/admin/roles/list", status_code=302)


# ========== 权限管理页面 ==========

# 权限列表页面
@router.get("/admin/permissions/list", response_class=HTMLResponse)
async def permissions_list(request: Request, page: int = 1, limit: int = 20, search: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import PermissionAdmin

    # 搜索功能
    query = PermissionAdmin.all()
    if search:
        query = query.filter(
            PermissionAdmin.name.contains(search) |
            PermissionAdmin.code.contains(search) |
            PermissionAdmin.category.contains(search)
        )

    offset = (page - 1) * limit
    permissions = await query.offset(offset).limit(limit).order_by("category", "name")
    total = await query.count()

    # 按分类分组显示
    permissions_by_category = {}
    for p in permissions:
        if p.category not in permissions_by_category:
            permissions_by_category[p.category] = []
        permissions_by_category[p.category].append(p)

    rows = ""
    for category, perms in permissions_by_category.items():
        rows += f'<tr class="category-row"><td colspan="6"><strong>{category}</strong></td></tr>'
        for p in perms:
            rows += f"""
            <tr>
                <td>{p.id}</td>
                <td>{p.name}</td>
                <td>{p.code}</td>
                <td>{p.category}</td>
                <td>{p.description or '-'}</td>
                <td>
                    <a href="/admin/permissions/edit/{p.id}" class="btn btn-sm">编辑</a>
                    <a href="/admin/permissions/delete/{p.id}" class="btn btn-sm btn-danger" onclick="return confirm('确定删除?')">删除</a>
                </td>
            </tr>
            """

    # 搜索表单
    search_form = f"""
    <div class="search-box">
        <form method="GET" action="/admin/permissions/list">
            <input type="text" name="search" placeholder="搜索权限名称、代码或分类" value="{search}">
            <button type="submit" class="btn btn-primary">搜索</button>
            <a href="/admin/permissions/list" class="btn">清空</a>
        </form>
    </div>
    """

    html = render_admin_page("权限管理", f"""
        <div class="page-header">
            <h2>权限管理</h2>
            {search_form}
            <a href="/admin/permissions/create" class="btn btn-primary">+ 新增权限</a>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>权限名称</th>
                    <th>权限代码</th>
                    <th>分类</th>
                    <th>描述</th>
                    <th>操作</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        <div class="pagination">
            <span>共 {total} 条记录</span>
        </div>
    """, user)
    return HTMLResponse(content=html)


# 创建权限页面
@router.get("/admin/permissions/create", response_class=HTMLResponse)
async def permission_create_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("创建权限", f"""
        <div class="page-header">
            <h2>创建权限</h2>
            <a href="/admin/permissions/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/permissions/create" class="data-form">
            <div class="form-group">
                <label>权限名称 *</label>
                <input type="text" name="name" required placeholder="例如: 基金管理">
            </div>
            <div class="form-group">
                <label>权限代码 *</label>
                <input type="text" name="code" required placeholder="例如: fund_manage">
            </div>
            <div class="form-group">
                <label>分类 *</label>
                <input type="text" name="category" required placeholder="例如: 系统管理">
            </div>
            <div class="form-group">
                <label>描述</label>
                <textarea name="description" rows="3" placeholder="权限描述"></textarea>
            </div>
            <button type="submit" class="btn btn-primary">提交</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 创建权限提交
@router.post("/admin/permissions/create")
async def permission_create_submit(
    name: str = Form(...),
    code: str = Form(...),
    category: str = Form(...),
    description: str = Form("")
):
    from admin_models import PermissionAdmin

    await PermissionAdmin.create(
        name=name,
        code=code,
        category=category,
        description=description
    )

    return RedirectResponse(url="/admin/permissions/list", status_code=302)


# 编辑权限页面
@router.get("/admin/permissions/edit/{permission_id}", response_class=HTMLResponse)
async def permission_edit_page(request: Request, permission_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import PermissionAdmin
    permission = await PermissionAdmin.filter(id=permission_id).first()

    if not permission:
        return HTMLResponse(content="<script>alert('权限不存在');history.back();</script>")

    html = render_admin_page("编辑权限", f"""
        <div class="page-header">
            <h2>编辑权限</h2>
            <a href="/admin/permissions/list" class="btn">返回列表</a>
        </div>
        <form method="POST" action="/admin/permissions/edit/{permission.id}" class="data-form">
            <div class="form-group">
                <label>权限名称 *</label>
                <input type="text" name="name" value="{permission.name}" required>
            </div>
            <div class="form-group">
                <label>权限代码 *</label>
                <input type="text" name="code" value="{permission.code}" required>
            </div>
            <div class="form-group">
                <label>分类 *</label>
                <input type="text" name="category" value="{permission.category}" required>
            </div>
            <div class="form-group">
                <label>描述</label>
                <textarea name="description" rows="3">{permission.description or ''}</textarea>
            </div>
            <button type="submit" class="btn btn-primary">保存修改</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 编辑权限提交
@router.post("/admin/permissions/edit/{permission_id}")
async def permission_edit_submit(
    permission_id: int,
    name: str = Form(...),
    code: str = Form(...),
    category: str = Form(...),
    description: str = Form("")
):
    from admin_models import PermissionAdmin
    permission = await PermissionAdmin.filter(id=permission_id).first()

    if not permission:
        return HTMLResponse(content="<script>alert('权限不存在');history.back();</script>")

    permission.name = name
    permission.code = code
    permission.category = category
    permission.description = description
    await permission.save()

    return RedirectResponse(url="/admin/permissions/list", status_code=302)


# 删除权限
@router.get("/admin/permissions/delete/{permission_id}")
async def permission_delete(request: Request, permission_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import PermissionAdmin, role_permissions
    permission = await PermissionAdmin.filter(id=permission_id).first()

    if not permission:
        return HTMLResponse(content="<script>alert('权限不存在');history.back();</script>")

    # 删除角色-权限关联
    await role_permissions.filter(permission_id=permission_id).delete()
    # 删除权限
    await permission.delete()

    return RedirectResponse(url="/admin/permissions/list", status_code=302)


# ========== 权限分配页面 ==========

# 权限分配页面
@router.get("/admin/roles/permissions/{role_id}", response_class=HTMLResponse)
async def role_permissions_page(request: Request, role_id: int):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import RoleAdmin, PermissionAdmin
    role = await RoleAdmin.filter(id=role_id).first()

    if not role:
        return HTMLResponse(content="<script>alert('角色不存在');history.back();</script>")

    # 获取所有权限
    all_permissions = await PermissionAdmin.all().order_by("category", "name")
    # 获取角色当前权限
    role_permissions = await role.permissions.all()
    role_permission_ids = [p.id for p in role_permissions]

    # 按分类分组
    permissions_by_category = {}
    for p in all_permissions:
        if p.category not in permissions_by_category:
            permissions_by_category[p.category] = []
        permissions_by_category[p.category].append(p)

    permission_checkboxes = ""
    for category, perms in permissions_by_category.items():
        permission_checkboxes += f'<div class="permission-category"><h4>{category}</h4>'
        for p in perms:
            checked = 'checked' if p.id in role_permission_ids else ''
            permission_checkboxes += f"""
            <div class="permission-item">
                <label>
                    <input type="checkbox" name="permissions" value="{p.id}" {checked}>
                    {p.name} ({p.code})
                    <span class="permission-desc">{p.description or ''}</span>
                </label>
            </div>
            """
        permission_checkboxes += '</div>'

    html = render_admin_page("权限分配", f"""
        <div class="page-header">
            <h2>权限分配 - {role.name}</h2>
            <a href="/admin/roles/list" class="btn">返回角色列表</a>
        </div>
        <form method="POST" action="/admin/roles/permissions/{role.id}" class="permission-form">
            <div class="permission-list">
                {permission_checkboxes}
            </div>
            <button type="submit" class="btn btn-primary">保存权限</button>
        </form>
    """, user)
    return HTMLResponse(content=html)


# 权限分配提交
@router.post("/admin/roles/permissions/{role_id}")
async def role_permissions_submit(
    role_id: int,
    permissions: list = Form(..., list=True)
):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import RoleAdmin, PermissionAdmin
    role = await RoleAdmin.filter(id=role_id).first()

    if not role:
        return HTMLResponse(content="<script>alert('角色不存在');history.back();</script>")

    # 获取当前权限
    current_permissions = await role.permissions.all()
    current_permission_ids = [p.id for p in current_permissions]

    # 添加新权限
    for perm_id in permissions:
        if int(perm_id) not in current_permission_ids:
            permission = await PermissionAdmin.filter(id=int(perm_id)).first()
            if permission:
                await role.permissions.add(permission)

    # 移除取消的权限
    for perm in current_permissions:
        if perm.id not in [int(p) for p in permissions]:
            await role.permissions.remove(perm)

    return RedirectResponse(url="/admin/roles/list", status_code=302)


# ========== 基金批量操作 ==========

# 批量删除
@router.post("/batch-delete")
async def funds_batch_delete(fund_ids: str = Form(...)):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    ids = fund_ids.split(",")
    await FundAdmin.filter(id__in=ids).delete()
    return RedirectResponse(url="/admin/funds/list", status_code=302)


# 批量更新状态
@router.post("/batch-update")
async def funds_batch_update(fund_ids: str = Form(...), status: str = Form(...)):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    ids = fund_ids.split(",")
    funds = await FundAdmin.filter(id__in=ids)
    for fund in funds:
        fund.status = status
        await fund.save()
    return RedirectResponse(url="/admin/funds/list", status_code=302)


# ========== 基金数据导出 ==========

# 导出Excel
@router.get("/export")
async def funds_export(request: Request, search: str = "", exchange: str = "", status: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    import io
    import csv
    from fastapi.responses import StreamingResponse

    # 查询数据
    query = FundAdmin.all()
    if search:
        query = query.filter(
            FundAdmin.fund_code.contains(search) |
            FundAdmin.fund_name.contains(search)
        )
    if exchange:
        query = query.filter(FundAdmin.exchange == exchange)
    if status:
        query = query.filter(FundAdmin.status == status)

    funds = await query

    # 创建CSV文件
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "基金代码", "基金名称", "交易所", "IPO日期", "IPO价格", "总份额", "净值", "管理人", "托管人", "资产类型", "状态"])

    for f in funds:
        writer.writerow([
            f.id, f.fund_code, f.fund_name, f.exchange or '',
            f.ipo_date or '', f.ipo_price or '', f.total_shares or '',
            f.nav or '', f.manager or '', f.custodian or '',
            f.asset_type or '', f.status or ''
        ])

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=funds_export.csv"}
    )


# 导出Excel格式
@router.get("/export/excel")
async def funds_export_excel(request: Request, search: str = "", exchange: str = "", status: str = ""):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    import io
    from fastapi.responses import StreamingResponse

    # 查询数据
    query = FundAdmin.all()
    if search:
        query = query.filter(
            FundAdmin.fund_code.contains(search) |
            FundAdmin.fund_name.contains(search)
        )
    if exchange:
        query = query.filter(FundAdmin.exchange == exchange)
    if status:
        query = query.filter(FundAdmin.status == status)

    funds = await query

    # 创建Excel文件
    output = io.BytesIO()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "基金数据"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # 写入表头
        headers = ["ID", "基金代码", "基金名称", "交易所", "IPO日期", "IPO价格", "总份额", "净值", "管理人", "托管人", "资产类型", "状态"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row, f in enumerate(funds, start=2):
            sheet.cell(row=row, column=1, value=f.id)
            sheet.cell(row=row, column=2, value=f.fund_code)
            sheet.cell(row=row, column=3, value=f.fund_name)
            sheet.cell(row=row, column=4, value=f.exchange or '')
            sheet.cell(row=row, column=5, value=f.ipo_date or '')
            sheet.cell(row=row, column=6, value=f.ipo_price or 0)
            sheet.cell(row=row, column=7, value=f.total_shares or 0)
            sheet.cell(row=row, column=8, value=f.nav or 0)
            sheet.cell(row=row, column=9, value=f.manager or '')
            sheet.cell(row=row, column=10, value=f.custodian or '')
            sheet.cell(row=row, column=11, value=f.asset_type or '')
            sheet.cell(row=row, column=12, value=f.status or '')

        # 设置列宽
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 12
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 15
        sheet.column_dimensions['K'].width = 15
        sheet.column_dimensions['L'].width = 12

        workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=funds_export.xlsx"}
        )
    except ImportError:
        # 如果openpyxl未安装，返回CSV格式
        import csv
        csv_output = io.StringIO()
        writer = csv.writer(csv_output)
        writer.writerow(headers)

        for f in funds:
            writer.writerow([
                f.id, f.fund_code, f.fund_name, f.exchange or '',
                f.ipo_date or '', f.ipo_price or 0, f.total_shares or 0,
                f.nav or 0, f.manager or '', f.custodian or '',
                f.asset_type or '', f.status or ''
            ])

        csv_output.seek(0)
        return StreamingResponse(
            csv_output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=funds_export.csv"}
        )


# 下载导入模板
@router.get("/template")
async def funds_template(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    import io
    import csv
    from fastapi.responses import StreamingResponse

    # 创建模板CSV文件
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["基金代码", "基金名称", "交易所", "IPO日期", "IPO价格", "总份额", "净值", "管理人", "托管人", "资产类型", "状态"])
    writer.writerow(["150050", "华夏中国交建REIT", "SH", "2021-12-20", "2.000", "1000000000", "2.150", "华夏基金", "建设银行", "基础设施", "listed"])
    writer.writerow(["示例数据，请替换为您的实际数据"])

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=funds_import_template.csv"}
    )


# ========== 基金数据导入 ==========

# 导入页面
@router.get("/import", response_class=HTMLResponse)
async def funds_import_page(request: Request):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    html = render_admin_page("基金数据导入", f"""
        <div class="page-header">
            <h2>基金数据导入</h2>
            <a href="/admin/funds/list" class="btn">返回列表</a>
        </div>
        <div class="import-box">
            <div class="import-instructions">
                <h3>导入说明</h3>
                <ol>
                    <li>请下载并填写基金导入模板</li>
                    <li>支持CSV格式文件</li>
                    <li>文件大小不超过10MB</li>
                    <li>基金代码不能重复</li>
                    <li>日期格式: YYYY-MM-DD</li>
                </ol>
                <div class="template-download">
                    <a href="/admin/funds/template" class="btn btn-primary">下载导入模板</a>
                </div>
            </div>
            <form method="POST" action="/admin/funds/import" class="import-form" enctype="multipart/form-data">
                <div class="form-group">
                    <label>选择文件</label>
                    <input type="file" name="file" accept=".csv" required>
                </div>
                <div class="form-group">
                    <label>导入模式</label>
                    <select name="mode">
                        <option value="append">追加模式 (保留现有数据)</option>
                        <option value="update">更新模式 (更新现有基金)</option>
                        <option value="replace">替换模式 (清空后导入)</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>
                        <input type="checkbox" name="validate" value="1" checked>
                        数据验证 (跳过错误数据)
                    </label>
                </div>
                <button type="submit" class="btn btn-primary">开始导入</button>
            </form>
        </div>
        <style>
            .import-box {{ max-width: 600px; margin: 0 auto; }}
            .import-instructions {{ background: #f8f9fa; padding: 20px; border-radius: 6px; margin-bottom: 20px; }}
            .import-instructions h3 {{ margin-bottom: 15px; }}
            .import-instructions ol {{ margin-left: 20px; margin-bottom: 15px; }}
            .import-instructions li {{ margin-bottom: 8px; }}
            .import-form {{ background: white; padding: 20px; border-radius: 6px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); }}
        </style>
    """, user)
    return HTMLResponse(content=html)


# 导入处理
@router.post("/import")
async def funds_import(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form("append"),
    validate: str = Form("1")
):
    user = get_admin_user(request)
    if not user:
        return RedirectResponse(url="/admin/login")

    from admin_models import FundAdmin
    import csv
    import io
    import pandas as pd

    # 读取文件内容
    contents = await file.read()

    # 根据文件类型选择读取方式
    if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
        # Excel文件处理
        try:
            df = pd.read_excel(io.BytesIO(contents))
            # 将DataFrame转换为字典列表
            df = df.fillna('')  # 处理空值
            reader = df.to_dict('records')

            # 检查必要的列是否存在
            required_columns = ["基金代码", "基金名称"]
            if not all(col in df.columns for col in required_columns):
                raise ValueError("Excel文件缺少必要的列: 基金代码, 基金名称")
        except Exception:
            logger.exception("Excel文件读取失败")
            return HTMLResponse(content="""
                <script>
                    alert("Excel文件读取失败，请检查文件格式");
                    window.location.href = "/admin/funds/import";
                </script>
            """)
    else:
        # CSV文件处理
        csv_file = io.StringIO(contents.decode('utf-8'))
        reader = csv.DictReader(csv_file)

    # 数据验证和导入
    success_count = 0
    error_count = 0
    errors = []

    if mode == "replace":
        await FundAdmin.all().delete()

    for row_num, row in enumerate(reader, start=2):
        try:
            # 验证必填字段
            if not row.get("基金代码") or not row.get("基金名称"):
                if validate == "1":
                    error_count += 1
                    continue
                else:
                    raise ValueError(f"第 {row_num} 行: 基金代码和基金名称不能为空")

            # 检查基金是否已存在
            fund_code = row["基金代码"]
            existing_fund = await FundAdmin.filter(fund_code=fund_code).first()

            fund_data = {
                "fund_code": fund_code,
                "fund_name": row.get("基金名称", ""),
                "exchange": row.get("交易所", ""),
                "ipo_date": row.get("IPO日期", ""),
                "ipo_price": float(row.get("IPO价格", 0)) if row.get("IPO价格") else None,
                "total_shares": float(row.get("总份额", 0)) if row.get("总份额") else None,
                "nav": float(row.get("净值", 0)) if row.get("净值") else None,
                "manager": row.get("管理人", ""),
                "custodian": row.get("托管人", ""),
                "asset_type": row.get("资产类型", ""),
                "status": row.get("状态", "pending"),
            }

            if mode == "update" and existing_fund:
                await existing_fund.update_from_dict(fund_data)
                await existing_fund.save()
            elif not existing_fund:
                await FundAdmin.create(**fund_data)
            else:
                if validate == "1":
                    error_count += 1
                    continue
                else:
                    raise ValueError(f"第 {row_num} 行: 基金代码 {fund_code} 已存在")

            success_count += 1

        except Exception as e:
            error_count += 1
            if validate != "1":
                errors.append(f"第 {row_num} 行: {str(e)}")

    # 返回导入结果
    result_message = f"导入完成: 成功 {success_count} 条, 失败 {error_count} 条"
    if errors:
        result_message += f"\\n错误信息:\\n" + "\\n".join(errors[:5])  # 只显示前5个错误

    return HTMLResponse(content=f"""
        <script>
            alert("{result_message}");
            window.location.href = "/admin/funds/list";
        </script>
    """)


# ========== 爬虫管理界面 ==========

# 爬虫管理主页面
